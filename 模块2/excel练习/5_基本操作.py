from openpyxl import workbook

wb = workbook.Workbook()  # Sheet

# 1. 修改sheet名称
"""
sheet = wb.worksheets[0]
sheet.title = "Python_excel"
wb.save("p3.xlsx")
"""

# 2. 创建sheet并设置sheet颜色
"""
sheet = wb.create_sheet("工作计划", 0)
sheet.sheet_properties.tabColor = "1072BA"
wb.save("p3.xlsx")
"""

# 3. 默认打开的sheet
"""
wb.active = 0
wb.save("p3.xlsx")
"""

# 4. 拷贝sheet
"""
sheet = wb.create_sheet("old")
sheet.sheet_properties.tabColor = "1072BA"

new_sheet = wb.copy_worksheet(wb["old"])
new_sheet.title = "new"
wb.save("p3.xlsx")
"""

# 5.删除sheet
"""
del wb["old"]
wb.save('p3.xlsx')
"""
