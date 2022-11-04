import os
import requests
from xml.etree import ElementTree
from openpyxl import workbook


# 1.提取XML格式中的数据
# 2.为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。
def main():
    wb = workbook.Workbook()
    while True:
        city = input('city:(N退出)>>>').strip()
        if city.upper() == 'N':
            break
        url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(
            city
        )
        res = requests.get(url=url)

        if 'Sheet' in wb.get_sheet_names():
            del wb['Sheet']
        tree = ElementTree.XML(res.content)
        sheet = wb.create_sheet(city)
        for index, option in enumerate(tree, 1):
            print(index, option.text)
            sheet.cell(index, 1).value = option.text
        wb.save(file_new_path)


if __name__ == '__main__':
    base_path = os.path.dirname(os.path.abspath(__file__))
    dir_path = os.path.join(base_path, 'files')
    file_path = os.path.join(dir_path, 'weather.xml')
    file_new_path = os.path.join(dir_path, 'weather.xlsx')
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    main()
