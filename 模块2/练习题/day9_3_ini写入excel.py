import os
from configparser import ConfigParser
from openpyxl import workbook
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill


def main():
    config = ConfigParser()
    config.read(file_path, encoding='utf-8')
    wb = workbook.Workbook()
    del wb['Sheet']
    my_side = Side(style="thin", color="000000")
    my_border = Border(top=my_side, bottom=my_side, left=my_side, right=my_side)
    my_font = Font(name="微软雅黑", size=45, color="FFFFFF", underline="single")
    my_fill = PatternFill("solid", fgColor="99ccff")
    my_alignment = Alignment(horizontal='center', vertical='distributed', wrap_text=False)
    for section in config.sections():
        sheet = wb.create_sheet(section)
        info_dic = {
            'A1': '键',
            'B1': '值',
        }
        for k, v in info_dic.items():
            cell = sheet[k]
            cell.value = v
            cell.font = my_font
            cell.fill = my_fill
            cell.border = my_border
            cell.alignment = my_alignment
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 35
        row_index = 2
        for items in config.items(section):
            for index, text in enumerate(items, 1):
                cell = sheet.cell(row_index, index)
                cell.alignment = my_alignment
                cell.border = my_border
                cell.value = text
            row_index += 1
        wb.save(file_new_path)


if __name__ == '__main__':
    base_path = os.path.dirname(os.path.abspath(__file__))
    dir_path = os.path.join(base_path, 'files')
    file_path = os.path.join(dir_path, 'my.ini')
    file_new_path = os.path.join(dir_path, 'my_ini_excel.xlsx')
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    main()
