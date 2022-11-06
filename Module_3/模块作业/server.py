"""
- 支持注册，并为用户初始化相关目录。
     注册成功之后，将所有用户信息存储到特定的Excel文件中
    用户信息的格式
        A,    B,    C
    1   用户名 密码 注册时间
    2
    3
- 支持登录
- 支持查看当前用户网盘目录下的所有文件。
- 支持上传
- 支持下载

请使用阻塞IO的方式实现
"""
import os
import json
import socket
from hashlib import md5

from openpyxl import load_workbook, workbook
from openpyxl.styles import Alignment, PatternFill

import setting
from utils import send_data, recv_data, recv_file, send_file
from res import Response


class Server:
    def __init__(self):
        self.ip = setting.IP
        self.port = setting.PORT
        self.server = socket.socket()
        self.server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.conn = None

        self.init_dir_path()

    def init_dir_path(self):
        if not os.path.exists(setting.USER_INFO_PATH):
            wb = workbook.Workbook()
            sheet = wb.worksheets[0]
            sheet.title = 'user_info'
            info_list = ['用户名', '密码', '注册时间']
            for index in range(len(info_list)):
                cell = sheet.cell(1, index + 1)
                cell.value = info_list[index]
                cell.alignment = Alignment(horizontal='center', vertical='distributed')
                cell.fill = PatternFill("solid", fgColor="00BFFF")
            wb.save(setting.USER_INFO_PATH)
        if not os.path.exists(setting.FILE_PATH):
            os.mkdir(setting.FILE_PATH)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.server:
            self.server.close()

    def run(self):
        self.server.bind((self.ip, self.port))
        self.server.listen(5)

        while True:
            self.conn, addr = self.server.accept()
            print('有连接来了...')
            while True:
                exec_dict = {
                    '1': self.register,
                    '2': self.login,
                    '3': self.upload_file,
                    '4': self.download_file,
                    '5': self.look_dir,
                }
                choice_num = recv_data(self.conn)
                if choice_num == 'close':
                    print('连接断开...')
                    break
                if choice_num == 'continue':
                    # print('continue')
                    continue
                params = recv_data(self.conn)
                if params == 'continue':
                    # print('continue')
                    continue
                # print(params)
                res = exec_dict[choice_num](params)
                send_data(self.conn, res)
            self.conn.close()

    def is_resgister(self, username):
        """
        判断用户是否已存在
        :param username:用户名
        :return:
        """
        wb = load_workbook(setting.USER_INFO_PATH)
        sheet = wb.worksheets[0]
        for cell in sheet:
            if cell[0].value == username:
                return True

    def enc_md5(self, enc_data):
        md5_obj = md5()
        md5_obj.update('梅花香自苦寒来'.encode('utf-8'))
        md5_obj.update(enc_data.encode('utf-8'))
        return md5_obj.hexdigest()

    def register(self, user_info):
        user_info = json.loads(user_info)
        username = user_info.get('1')

        is_reg = self.is_resgister(username)
        if is_reg:
            res_data = Response()
            res_data.code = 400
            res_data.message = '用户已存在!'
            return res_data.get_info()

        user_dir = os.path.join(setting.FILE_PATH, username)
        os.mkdir(user_dir)

        pwd = user_info.get('2')
        new_pwd = self.enc_md5(pwd)
        user_info.update({'2': new_pwd})

        wb = load_workbook(setting.USER_INFO_PATH)
        sheet = wb.worksheets[0]
        current_row = sheet.max_row + 1
        for i in range(1, 4):
            cell = sheet.cell(current_row, i)
            cell.value = user_info.get('%s' % i)
        wb.save(setting.USER_INFO_PATH)
        print(f"{user_info.get('1')}---注册成功")
        res_data = Response(message=f'{username}用户注册成功').get_info()
        return res_data

    def login(self, user_info):
        user_info = json.loads(user_info)
        username = user_info['username']
        password = self.enc_md5(user_info['password'])
        wb = load_workbook(setting.USER_INFO_PATH)
        sheet = wb.worksheets[0]
        for cell in sheet:
            if cell[0].value == username and cell[1].value == password:
                message = f'{username}---登录成功!'
                print(message)
                return Response(message=message).get_info()
        return Response(code=400, message='用户名或密码错误!').get_info()

    def upload_file(self, file_name):
        recv_file(self.conn, file_name)
        res_data = Response(message='上传成功!').get_info()
        return res_data

    def download_file(self, file_name, current_seek=0):
        user_path = os.path.join(setting.FILE_PATH, file_name)
        if not os.path.exists(user_path):
            return Response(code=400, message='网盘无此文件!').get_info()
        if current_seek == 0:
            # 无断点续传
            send_data(self.conn, Response(message='开始下载文件!').get_info())

            send_file(self.conn, user_path)
            return Response(message='文件下载成功!').get_info()
        else:
            print('开始续传文件')
            send_data(self.conn, Response(message='开始下载文件(续传)!').get_info())

            print('开始发送续传文件')
            send_file(self.conn, user_path, current_seek)
            return Response(message='文件下载(续传)成功!').get_info()

    def look_dir(self, cmd):
        if cmd != 'ls':
            return Response(code=400, message='命令错误!').get_info()
        username = recv_data(self.conn)
        # 用户网盘路径
        user_dirs = os.path.join(setting.FILE_PATH, username)
        data = os.listdir(user_dirs)
        return Response(data=data).get_info()


if __name__ == '__main__':
    with Server() as server_obj:
        server_obj.run()
