from openpyxl import load_workbook

"""源文件修改"""
# wb = load_workbook('p1.xlsx')
# sheet = wb.worksheets[0]
#
# # 找到单元格，并修改单元格的内容
# cell = sheet.cell(1, 1)
# cell.value = "新的开始"
#
# # 将excel文件保存到p2.xlsx文件中
# wb.save("p2.xlsx")


"""新创建文件"""
# from openpyxl import workbook
#
# # 创建excel且默认会创建一个sheet（名称为Sheet）
# wb = workbook.Workbook()
#
# sheet = wb.worksheets[0]  # 或 sheet = wb["Sheet"]
#
# # 找到单元格，并修改单元格的内容
# cell = sheet.cell(1, 1)
# cell.value = "workbook.Workbook"
#
# # 将excel文件保存到p2.xlsx文件中
# wb.save("p3.xlsx")


