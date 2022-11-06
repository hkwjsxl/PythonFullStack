from openpyxl import load_workbook

wb = load_workbook("p1.xlsx")
sheet = wb.worksheets[2]

# 获取第N行第N列的单元格(位置是从1开始）
c1 = sheet.cell(1, 1)
print(c1)  # <Cell 'Sheet1'.A1>
print(c1.value)  # 用户信息

c2 = sheet.cell(1, 2)
print(c2)  # <MergedCell 'Sheet1'.B1>
print(c2.value)  # None

""""""

from openpyxl import load_workbook

wb = load_workbook('p1.xlsx')
sheet = wb.worksheets[2]
for row in sheet.rows:
    print(row)
