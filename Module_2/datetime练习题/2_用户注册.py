"""用户注册，将用户信息写入Excel，其中包含：用户名、密码、注册时间 三列"""
import os
from hashlib import md5
from datetime import datetime

from openpyxl import load_workbook, workbook


def enc_md5(enc_data):
    md5_obj = md5()
    md5_obj.update(enc_data.encode('utf-8'))
    return md5_obj.hexdigest()


def register(username, password):
    password = enc_md5(password)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.worksheets[0]
        next_row = sheet.max_row + 1
    else:
        wb = workbook.Workbook()
        sheet = wb.worksheets[0]
        next_row = 1
    sheet.cell(next_row, 1).value = username
    sheet.cell(next_row, 2).value = password
    sheet.cell(next_row, 3).value = datetime.now().strftime('%Y-%m-%d')
    if sheet.title != 'user_info':
        sheet.title = 'user_info'
    wb.save(file_path)


def main():
    while True:
        user = input('请输入用户名(Q):').strip()
        if user.upper() == 'Q':
            break
        pwd = input('请输入密码:').strip()
        register(user, pwd)


if __name__ == '__main__':
    base_path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_path, 'user_info.xlsx')
    main()
